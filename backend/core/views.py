# 🔧 Utilidades estándar
import json
import logging
from decimal import Decimal, InvalidOperation
from datetime import datetime
from functools import wraps

# 📦 Librerías externas
import openpyxl

# 🌐 Django - HTTP y vistas
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt

# 🔐 Django - Autenticación
from django.contrib.auth import authenticate, login, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.views import LogoutView

# 🧠 Django - Utilidades
from django.utils import timezone
from django.db.models import Sum
from django.contrib import messages

# 🗂️ Modelos y formularios locales
from .models import (
    Platillo,
    PerfilUsuario,
    Mesa,
    Cuenta,
    Orden,
    GastoExtra,
    CorteCaja
)
from .forms import RegistroUsuarioForm, EditarUsuarioForm

# 📝 Logger
logger = logging.getLogger(__name__)

# -------------------------
# Decoradores
# -------------------------
def solo_admin(view_func):
    """Restringe el acceso solo a usuarios con role='admin'."""
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        perfil = getattr(request.user, 'perfilusuario', None)
        if request.user.is_authenticated and perfil and perfil.role == 'admin':
            return view_func(request, *args, **kwargs)
        messages.error(request, '⛔ No tienes permisos para realizar esta acción.')
        return redirect('menu')
    return _wrapped_view

# -------------------------
# Autenticación
# -------------------------
def login_view(request):
    if request.method == 'POST':
        input_usuario = request.POST.get('username')
        password = request.POST.get('password')

        # Soporte login por email
        if '@' in input_usuario:
            user_obj = User.objects.filter(email=input_usuario).first()
            if not user_obj:
                return render(request, 'inicioSesion.html', {'error': 'Correo no registrado'})
            username = user_obj.username
        else:
            username = input_usuario

        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            perfil = getattr(user, 'perfilusuario', None)
            request.session['user_role'] = perfil.role if perfil else 'employee'
            return redirect('menu')
        return render(request, 'inicioSesion.html', {'error': 'Credenciales inválidas'})

    return render(request, 'inicioSesion.html')


def nuevo_usuario(request):
    form = RegistroUsuarioForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        perfil = getattr(user, 'perfilusuario', None)
        request.session['user_role'] = perfil.role if perfil else 'employee'
        return redirect('login')
    return render(request, 'nuevo_usuario.html', {'form': form})


class LogoutConMensaje(LogoutView):
    """Logout con mensaje de éxito."""
    def dispatch(self, request, *args, **kwargs):
        messages.success(request, "Sesión finalizada correctamente. Gracias por utilizar Cuenta Clara.")
        return super().dispatch(request, *args, **kwargs)

# -------------------------
# Vistas generales
# -------------------------
@login_required
def menu(request):
    return render(request, 'menu.html')

@login_required
def ajustes(request):
    return render(request, 'ajustes.html')

@login_required
def menu_comida(request):
    platillos = Platillo.objects.filter(activo=True)
    return render(request, 'menu_comida.html', {'platillos': platillos})

@login_required
def mesas(request):
    mesas = Mesa.objects.order_by('numero')
    # Datos temporales para ejemplo
    ordenes = [
        {'id': 1, 'nombre_mesa': 'Mesa 3', 'items': ['2x Hamburguesa Clásica'], 'para_llevar': False},
        {'id': 2, 'nombre_mesa': 'Para Llevar', 'items': ['3x Tacos al Pastor'], 'para_llevar': True}
    ]
    return render(request, 'mesas.html', {'mesas': mesas, 'ordenes': ordenes})

@login_required
def cuentas_view(request):
    mesa_filtro = request.GET.get('mesa')
    fecha_filtro = request.GET.get('fecha')

    cuentas = Cuenta.objects.prefetch_related('ordenes__platillos').order_by('-creada')
    if mesa_filtro:
        cuentas = cuentas.filter(mesa__numero=mesa_filtro)
    if fecha_filtro:
        cuentas = cuentas.filter(creada__date=fecha_filtro)

    return render(request, 'cuentas.html', {'cuentas': cuentas})


@login_required
def cerrar_cuenta(request, cuenta_id):
    cuenta = get_object_or_404(Cuenta, id=cuenta_id, activa=True)
    if request.method == 'POST':
        cuenta.activa = False
        cuenta.calcular_total()
        cuenta.save()
        messages.success(request, f'Cuenta de mesa {cuenta.mesa.numero} cerrada correctamente.')
        return redirect('cuentas')
    return render(request, 'cuentas.html', {'cuentas': Cuenta.objects.filter(activa=True)})

# -------------------------
# Gestión de usuarios
# -------------------------
@login_required
@solo_admin
def centro_de_usuarios(request):
    rol = request.GET.get('rol')
    usuarios = User.objects.select_related('perfilusuario')
    if rol:
        usuarios = usuarios.filter(perfilusuario__role=rol)
    return render(request, 'Ajustes/centro_de_usuarios.html', {'usuarios': usuarios, 'rol_seleccionado': rol})

@login_required
@solo_admin
def agregar_usuario(request):
    form = RegistroUsuarioForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        PerfilUsuario.objects.get_or_create(user=user, defaults={'role': 'employee'})
        messages.success(request, '✅ Usuario agregado correctamente')
        return redirect('centro_de_usuarios')
    return render(request, 'Ajustes/agregar_usuario.html', {'form': form})

@login_required
@solo_admin
def editar_usuario(request, usuario_id):
    usuario = get_object_or_404(User, id=usuario_id)
    if not request.user.is_superuser and request.user != usuario:
        messages.error(request, '❌ No tienes permiso para editar este usuario.')
        return redirect('centro_de_usuarios')

    form = EditarUsuarioForm(request.POST or None, request.FILES or None, instance=usuario)
    if request.method == 'POST' and form.is_valid():
        form.save()
        if request.user == usuario:
            update_session_auth_hash(request, usuario)
        messages.success(request, '✅ Datos actualizados correctamente.')
        return redirect('centro_de_usuarios')
    elif request.method == 'POST':
        messages.error(request, '❌ Corrige los errores del formulario.')

    return render(request, 'Ajustes/editar_usuario.html', {'form': form, 'usuario': usuario})

@login_required
@solo_admin
def eliminar_usuario(request, usuario_id):
    usuario_a_eliminar = get_object_or_404(User, id=usuario_id)
    if usuario_a_eliminar == request.user:
        messages.error(request, '⛔ No puedes eliminar tu propio usuario.')
        return redirect('centro_de_usuarios')

    perfil = getattr(usuario_a_eliminar, 'perfilusuario', None)
    if perfil and perfil.role == 'admin' and User.objects.filter(perfilusuario__role='admin').count() <= 1:
        messages.error(request, '⛔ No puedes eliminar al único administrador.')
        return redirect('centro_de_usuarios')

    usuario_a_eliminar.delete()
    messages.success(request, '🗑️ Usuario eliminado correctamente.')
    return redirect('centro_de_usuarios')

# -------------------------
# Gestión de platillos
# -------------------------
@login_required
@solo_admin
def editar_menu(request):
    platillos = Platillo.objects.all()
    return render(request, 'Ajustes/editar_menu.html', {'platillos': platillos})

@login_required
@solo_admin
def agregar_platillo(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        precio_raw = request.POST.get('precio', '').strip()
        ingredientes = request.POST.getlist('ingredientes')
        foto = request.FILES.get('foto')

        try:
            precio = Decimal(precio_raw)
            if precio < 0:
                raise InvalidOperation("Precio negativo")
        except (InvalidOperation, TypeError, ValueError):
            messages.error(request, '❌ El precio ingresado no es válido.')
            return redirect('agregar_platillo')

        ingredientes_str = ', '.join([i.strip() for i in ingredientes if i.strip()])
        Platillo.objects.create(user=request.user, nombre=nombre, precio=precio, ingredientes=ingredientes_str, foto=foto)
        messages.success(request, '✅ Platillo agregado correctamente.')
        return redirect('editar_menu')

    return render(request, 'Ajustes/agregar_platillo.html')

@login_required
@solo_admin
def editar_platillo(request, platillo_id):
    platillo = get_object_or_404(Platillo, id=platillo_id)

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        precio_raw = request.POST.get('precio', '').strip()
        ingredientes = request.POST.getlist('ingredientes[]')

        if not nombre or not precio_raw:
            messages.error(request, '❌ Nombre y precio obligatorios.')
            return redirect('editar_platillo', platillo_id=platillo.id)

        try:
            precio = Decimal(precio_raw)
        except InvalidOperation:
            messages.error(request, '❌ Precio inválido.')
            return redirect('editar_platillo', platillo_id=platillo.id)

        platillo.nombre = nombre
        platillo.precio = precio
        platillo.ingredientes = ', '.join([i.strip() for i in ingredientes if i.strip()])
        if request.FILES.get('foto'):
            platillo.foto = request.FILES['foto']
        platillo.save()

        messages.success(request, '✅ Platillo actualizado correctamente.')
        return redirect('editar_menu')

    ingredientes_list = [i.strip() for i in platillo.ingredientes.split(',')] if platillo.ingredientes else []
    campos_vacios = range(max(0, 10 - len(ingredientes_list)))

    return render(request, 'Ajustes/editar_platillo.html', {
        'platillo': platillo,
        'ingredientes_list': ingredientes_list,
        'campos_vacios': campos_vacios
    })

@login_required
@solo_admin
def eliminar_platillo(request, platillo_id):
    platillo = get_object_or_404(Platillo, id=platillo_id)
    platillo.delete()
    messages.success(request, '🗑️ Platillo eliminado correctamente.')
    return redirect('editar_menu')

@login_required
@solo_admin
def actualizar_activo(request, id):
    platillo = get_object_or_404(Platillo, id=id)
    platillo.activo = not platillo.activo
    platillo.save()
    messages.success(request, f"Estado de '{platillo.nombre}' actualizado.")
    return redirect('editar_menu')

# -------------------------
# Pedidos y órdenes
# -------------------------
@login_required
def procesar_pedido(request):
    if request.method != 'POST':
        messages.error(request, '⛔ Método no permitido.')
        return redirect('menu_comida')

    ids_str = request.POST.get('platillos[]', '')
    ids = ids_str.split(',') if ids_str else []
    if not ids:
        messages.error(request, '⚠️ No seleccionaste ningún platillo.')
        return redirect('menu_comida')

    messages.success(request, '✅ Pedido enviado correctamente.')
    return redirect('menu_comida')

@login_required
def crear_orden(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("Método no permitido")

    try:
        platillos_data = request.POST.get("platillos_seleccionados")
        mesa_id = request.POST.get("mesa")

        if not platillos_data:
            return HttpResponseBadRequest("No se recibieron platillos")

        platillo_ids = json.loads(platillos_data)
        if not platillo_ids:
            return HttpResponseBadRequest("Debes seleccionar al menos un platillo")

        # Buscar o crear cuenta activa
        cuenta = None
        if mesa_id:
            try:
                mesa = Mesa.objects.get(id=mesa_id)
            except Mesa.DoesNotExist:
                return HttpResponseBadRequest("Mesa no encontrada")
            cuenta, _ = Cuenta.objects.get_or_create(mesa=mesa, activa=True, defaults={"usuario": request.user})
        else:
            cuenta, _ = Cuenta.objects.get_or_create(mesa=None, activa=True, usuario=request.user)

        # Crear orden
        orden = Orden.objects.create(cuenta=cuenta, usuario=request.user)
        orden.platillos.set(Platillo.objects.filter(id__in=platillo_ids))
        orden.save()

        cuenta.platillos.add(*orden.platillos.all())
        cuenta.calcular_total()
        return redirect("cuentas")

    except json.JSONDecodeError:
        return HttpResponseBadRequest("Error en los datos de platillos")
    except Exception as e:
        logger.error(f"Error al crear orden: {e}")
        return HttpResponseBadRequest(f"Error inesperado: {str(e)}")

# -------------------------
# AJAX
# -------------------------
@csrf_exempt
def actualizar_pedido(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)
    try:
        data = json.loads(request.body)
        return JsonResponse({"success": True, "data": data})
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)

# -------------------------
# Gestión de mesas
# -------------------------
@login_required
@solo_admin
def editar_mesas(request):
    mesas = Mesa.objects.order_by('numero')
    return render(request, 'Ajustes/editar_mesas.html', {'mesas': mesas})

@login_required
@solo_admin
def agregar_mesa(request):
    if request.method == 'POST':
        numero = request.POST.get('numero')
        color = request.POST.get('color')
        if not numero or Mesa.objects.filter(numero=numero).exists():
            messages.error(request, '⚠️ Número inválido o ya existe.')
            return redirect('agregar_mesa')
        Mesa.objects.create(numero=numero, color=color)
        return redirect('editar_mesas')
    return render(request, 'Ajustes/agregar_mesa.html')

@require_POST
@login_required
def eliminar_mesa(request, numero):
    try:
        mesa = Mesa.objects.get(numero=numero)
        mesa.delete()
        return JsonResponse({'success': True})
    except Mesa.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Mesa no encontrada'})

# -------------------------
# Corte de caja
# -------------------------
def vista_corte(request):
    # Obtener fecha desde POST o usar la actual
    fecha_str = request.POST.get("fecha")
    fecha = timezone.now().date()
    if fecha_str:
        try:
            fecha = timezone.datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            pass  # Si la fecha es inválida, se mantiene la actual

    # Datos base del corte
    cuentas_cerradas = Cuenta.objects.filter(cerrada__date=fecha)
    ventas_totales = cuentas_cerradas.aggregate(total=Sum('total'))['total'] or 0

    gastos = GastoExtra.objects.filter(fecha=fecha)
    gastos_totales = gastos.aggregate(total=Sum('monto'))['total'] or 0

    corte_existente = CorteCaja.objects.filter(fecha=fecha).first()
    monto_extra = corte_existente.monto_extra if corte_existente else 0
    dinero_en_caja = (
        corte_existente.dinero_en_caja if corte_existente
        else ventas_totales - gastos_totales + monto_extra
    )

    mensaje = ""

    # Agregar monto extra
    if request.method == "POST" and "pagos_extra" in request.POST:
        try:
            monto_extra = Decimal(request.POST.get("monto_extra", "0"))
            mensaje = "Monto extra agregado correctamente."
        except (InvalidOperation, TypeError):
            monto_extra = 0
            mensaje = "Monto extra inválido."

    # Calcular corte
    if request.method == "POST" and "calcular_corte" in request.POST:
        try:
            efectivo_inicial = Decimal(request.POST.get("efectivo_inicial", "0"))
        except (InvalidOperation, TypeError):
            efectivo_inicial = 0

        dinero_en_caja = efectivo_inicial + ventas_totales - gastos_totales + monto_extra

        CorteCaja.objects.update_or_create(
            fecha=fecha,
            defaults={
                "efectivo_inicial": efectivo_inicial,
                "ventas_totales": ventas_totales,
                "gastos_totales": gastos_totales,
                "monto_extra": monto_extra,
                "dinero_en_caja": dinero_en_caja,
                "creado_por": request.user
            }
        )
        mensaje = "Corte de caja guardado correctamente."

    context = {
        "fecha": fecha,
        "ventas_totales": ventas_totales,
        "gastos_totales": gastos_totales,
        "monto_extra": monto_extra,
        "dinero_en_caja": dinero_en_caja,
        "mensaje": mensaje,
    }
    return render(request, "corte.html", context)


def eliminar_cuenta(request, cuenta_id):
    cuenta = get_object_or_404(Cuenta, id=cuenta_id)
    cuenta.delete()
    messages.success(request, "Cuenta eliminada correctamente.")
    return redirect('cuentas')


def exportar_corte_excel(request):
    fecha_str = request.GET.get("fecha")
    fecha = timezone.now().date()
    if fecha_str:
        fecha = timezone.datetime.strptime(fecha_str, "%Y-%m-%d").date()

    corte = CorteCaja.objects.filter(fecha=fecha).first()
    gastos = GastoExtra.objects.filter(fecha=fecha)
    cuentas = Cuenta.objects.filter(cerrada__date=fecha)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corte de Caja"

    # Encabezados
    ws.append(["Fecha", "Efectivo Inicial", "Ventas Totales", "Gastos Totales", "Monto Extra", "Dinero en Caja"])
    if corte:
        ws.append([
            corte.fecha.strftime("%d/%m/%Y"),
            float(corte.efectivo_inicial),
            float(corte.ventas_totales),
            float(corte.gastos_totales),
            float(corte.monto_extra),
            float(corte.dinero_en_caja)
        ])
    else:
        ws.append([fecha.strftime("%d/%m/%Y"), "", "", "", "", ""])

    # Detalle de gastos
    ws_gastos = wb.create_sheet(title="Gastos")
    ws_gastos.append(["Fecha", "Descripción", "Monto"])
    for gasto in gastos:
        ws_gastos.append([
            gasto.fecha.strftime("%d/%m/%Y"),
            gasto.descripcion,
            float(gasto.monto)
        ])

    # Detalle de cuentas
    ws_cuentas = wb.create_sheet(title="Cuentas Cerradas")
    ws_cuentas.append(["Mesa", "Total", "Fecha de cierre"])
    for cuenta in cuentas:
        mesa = cuenta.mesa.numero if cuenta.mesa else "Para llevar"
        ws_cuentas.append([
            mesa,
            float(cuenta.total),
            cuenta.cerrada.strftime("%d/%m/%Y %H:%M")
        ])

    # Respuesta HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"Corte_{fecha.strftime('%Y%m%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
